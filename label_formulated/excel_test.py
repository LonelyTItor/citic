from openpyxl import Workbook,load_workbook
from openpyxl.styles import *
import numpy as np

if __name__ == '__main__':
    # load tag lists
    wb = load_workbook('./data/tempTags.xlsx')
    ws = wb['Sheet1']
    tag_dict = {}
    for i in range(2, 67060):
        word = ws.cell(i, 1).value
        tag_dict[word] = np.zeros(300, dtype=np.float32)
    print(tag_dict)