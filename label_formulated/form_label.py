from gensim.models.word2vec import KeyedVectors
from openpyxl import Workbook,load_workbook
from openpyxl.styles import *
import numpy as np
import jiagu
import jieba
import thulac
from sklearn.cluster import KMeans
thu1 = thulac.thulac()

NUM_TAG = 67058
NUM_CLUSTERS = 1000

def split_match(miss_word, wv):
    word_1 = jiagu.seg(miss_word)
    word_2 = jieba.lcut(miss_word)
    word_3 = thu1.cut(miss_word)

    # select the shortest one
    length = np.array([len(word_1), len(word_2), len(word_3)])
    id = np.argmin(length)
    ready_word = eval('word_{}'.format(id+1))
    ready_len = len(ready_word)

    # then put it into the dict
    temp = np.zeros(200, dtype=np.float32)
    for i in range(ready_len):
        if ready_word[i] == '':
            ready_len -=1
            continue
        try:
            temp += wv[ready_word[i]]
        except:
            # give up one word
            ready_len -=1
            continue

    if ready_len == 0:
        print('split match failed')
        return np.zeros(200, dtype=np.float32), 0
    # method for process the vector
    temp /= ready_len

    # try to find similars, and the second one is better
    #sim_a = wv.most_similar(positive=ready_word, topn=5)
    #sim_b = wv.similar_by_vector(vector=temp, topn=5)

    print('split match success', ready_word)

    return temp, 1


def get_list(my_ws, wv):
    tag_dict = np.zeros((NUM_TAG, 200), dtype=np.float32)
    count = 0
    for i in range(2, NUM_TAG):
        word = my_ws.cell(i, 1).value
        # the qurey should be write down here.
        try:
            tag_dict[i-2] = wv[word]
            print("{} success".format(word))
            count += 1
        except:
            print("{} not found, try split match".format(word))
            tag_dict[i-2], val = split_match(word, wv)
            count += val
            continue
    print(count)
    return tag_dict

def get_wv():
    file = '/data/wordModel/Tencent_AILab_ChineseEmbedding.txt'
    wv = KeyedVectors.load_word2vec_format(file, binary=False)
    return wv


def visualize(wv, mapping_word, center, labels, ws):
    first_list = []
    second_list = []
    for idx in range(NUM_CLUSTERS):
        cur_vec = center[idx]
        sims = wv.similar_by_vector(vector=cur_vec, topn=2)
        first_list.append(sims[0])
        second_list.append(sims[1])
        print(sims, '------', idx)
    for jdx in range(mapping_word.shape[0]):
        ans1 = first_list[labels[jdx]]
        ans2 = second_list[labels[jdx]]
        real_jdx = mapping_word[jdx]
        ws.cell(real_jdx+2, 2).value = ans1[0]
        ws.cell(real_jdx+2, 3).value = ans2[0]
    return

init_list ='./data/w2v_final.npy'
if __name__ == '__main__':

    if init_list is None:
        # load word2vector Model
        wv_from_text = get_wv()
        print('aaaaa')

        # load tag lists

        wb = load_workbook('./data/tempTags.xlsx')
        ws = wb['Sheet1']
        # initial the tag dict
        aa = get_list(ws, wv_from_text)
        np.save('./data/w2v_final', aa)
        print('aaaa')
    else:
        aa = np.load(init_list)

    # here we come to start clustering

    # mask the zeros first
    zero_flag = np.sum(aa, axis=1)
    X_data = aa[zero_flag != 0, :]
    aux_mapping = np.ones(NUM_TAG, dtype=np.int32) * -1
    for i in range(NUM_TAG):
        aux_mapping[i] = i
    mapping_word = aux_mapping[zero_flag != 0]

    ## start clustering
    # kmeans= KMeans(n_clusters=NUM_CLUSTERS, max_iter=1000).fit(X_data)

    ## start visualization
    wv = get_wv()
    center = np.load('./data/kmeans_1000/centers.npy')
    labels = np.load('./data/kmeans_1000/label_classification.npy')
    wb = load_workbook('./data/tempTags.xlsx')
    ws = wb['Sheet1']
    visualize(wv, mapping_word, center, labels, ws)
    wb.save('./data/out.xlsx')
    print('bbbb')
